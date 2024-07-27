from pathlib import Path
from openpyxl import load_workbook
from collections import namedtuple
from datetime import date

WORKBOOK = 'check.xlsx'
SHEET = 'ЖурналПоверки'
MNF_NUMBER = 4
FIF_NUMBER = 5
TARGET = 6
DATE_DISPATCH = 10
APPLICABILITY = 14
VERIFICATION_DATE = 15
DOCNUM = 16
DATE_FORMAT = 'DD.MM.YYYY'


Mi = namedtuple('Mi', 'number fif date_dispatch')
Result = namedtuple('Result', 'applicability docnum date')
ItemFields = namedtuple('ItemFields', 'number fif date_dispatch')
ResultFields = namedtuple('ResultFields', 'applicability date docnum')
item_fields = ItemFields(MNF_NUMBER, FIF_NUMBER, DATE_DISPATCH)
result_fields = ResultFields(APPLICABILITY, VERIFICATION_DATE, DOCNUM)

file = Path.cwd() / WORKBOOK


def get_data(file):
    data = {}
    wb = load_workbook(file, data_only=True)
    sheet = wb[SHEET]

    for row in range(2, sheet.max_row + 1):
        if sheet.cell(row=row, column=DOCNUM).value:
            continue
        if sheet.cell(row=row, column=TARGET).value != 'Поверка':
            continue
        data[row] = {
            'mi': Mi(*get_mi(sheet, row, item_fields)),
            'result': Result(*get_result(sheet, row, result_fields))
        }

    wb.close()
    return data


def set_data(file, data):
    wb = load_workbook(file, data_only=True)
    sheet = wb[SHEET]
    for key, value in data.items():
        if value['result'].docnum:
            set_result(sheet, key, value['result'])
    wb.save(file)
    wb.close()


def get_mi(sheet, row, fields):
    return (
        sheet.cell(row=row, column=fields.number).value,
        sheet.cell(row=row, column=fields.fif).value,
        sheet.cell(row=row, column=fields.date_dispatch).value
    )

def get_result(sheet, row, fields=result_fields):
    return (
        sheet.cell(row=row, column=fields.applicability).value,
        sheet.cell(row=row, column=fields.date).value,
        sheet.cell(row=row, column=fields.docnum).value
    )


def set_result(sheet, row, item, fields=result_fields):
    sheet.cell(row=row, column=fields.applicability).value = item.applicability
    sheet.cell(row=row, column=fields.docnum).value = item.docnum
    sheet.cell(row=row, column=fields.date).value = item.date
    sheet.cell(row=row, column=fields.date).number_format = DATE_FORMAT



def insert_result(item, result):
    item['result'] = Result(**result)
    return item

if __name__ == '__main__':
    print('=======================')
    print(get_data(file))
    # rrr = {
    #     'applicability': 'свидетельство',
    #     'docnum': 'С-АУ/18-07-2024/355476803',
    #     'date': date.today()
    # }
    # sss = {
    #     'applicability': 'извещение',
    #     'docnum': 'И-АУ/14-07-2024/1111111111',
    #     'date': date.today()
    # }
    # data = get_data(file)
    # print(data)
    # print('=======================')
    # data[2]= insert_result(data[2],rrr)
    # data[4]= insert_result(data[4],sss)
    # print(data)
    # set_data(file, data)