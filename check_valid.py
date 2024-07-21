import requests
from pathlib import Path
from time import sleep
from openpyxl import load_workbook
import datetime


# Адрес API ГИС Аршин
URL = 'https://fgis.gost.ru/fundmetrology/eapi/vri?'
WORKBOOK = 'check.xlsx'
SHEET = 'ЖурналПоверки'
DATE_URL = '%Y-%m-%d'
DATE_XLS = '%d.%m.%Y'


# Отбрасываем СИ поверенные до дату отправки на поверку
def get_actual_date(response, date_of_dispatch):
    response_temp = []
    for mi in response:
        if date_of_dispatch > datetime.datetime.strptime(mi['verification_date'], DATE_XLS):
            continue
        response_temp.append(mi)
    return response_temp


# Формируем ссылку для проверки результата по зав.номеру и ФИФ
def get_url(mi_number, mit_number, verification_date_start: datetime.datetime):
    url_mi_number = f'mi_number={mi_number}'
    url_mit_number = f'mit_number={mit_number}'
    url_verification_date_start = f'verification_date_start={verification_date_start.strftime(DATE_URL)}'
    url = URL + url_mi_number + '&' + url_mit_number + '&' + url_verification_date_start
    return url


# Добавляем в словарь результаты поверки
def add_result(response, mi):
    mi['mit_number'] = response[0]['mit_number']
    mi['mi_number'] = response[0]['mi_number']
    mi['verification_date'] = response[0]['verification_date']
    mi['result_docnum'] = response[0]['result_docnum']
    return mi


def xldate(xldate, datemode=0):
    # datemode: 0 for 1900-based, 1 for 1904-based
    return (
        datetime.datetime(1899, 12, 30)
        + datetime.timedelta(days=xldate + 1462 * datemode)
        )


# Путь до файла со списком СИ
path = Path(Path.cwd())
file = path / WORKBOOK

# Открываем лист с таблицей из файла со списком СИ
wb = load_workbook(file, data_only=True)
sheet = wb[SHEET]

# Читаем данные из таблицы
for row in range(2, sheet.max_row + 1):

    # Отбираем СИ не полученные от других организаций
    if sheet.cell(row=row, column=12).value is not None:
        # print(row, '===', 12, '===', sheet.cell(row=row, column=12).value)
        continue
    # Отбираем СИ переданные в другие организации
    if sheet.cell(row=row, column=10).value is None:
        # print(row, '===', 10, '===', sheet.cell(row=row, column=10).value)
        continue
    # Отбираем СИ находящиеся на поверке
    if sheet.cell(row=row, column=6).value!='Поверка':
        # print(row, '===', 6, '===', sheet.cell(row=row, column=6).value)
        continue

    # Отбираем СИ прошедние процедуру проверки
    if sheet.cell(row=row, column=14).value is not None:
        # print(row, '===', 14, '===', sheet.cell(row=row, column=14).value)
        continue


    mi_number = sheet.cell(row=row, column=4).value
    mit_number = sheet.cell(row=row, column=5).value
    date_of_dispatch = sheet.cell(row=row, column=10).value
    
    # Создать ссылку для проверки
    url = get_url(
        mi_number=mi_number,
        mit_number=mit_number,
        verification_date_start=date_of_dispatch
        )

    # Отравляем запрос
    try:
        response = requests.get(url).json()['result']['items']
    except:
        print('По придержи коней')
        sleep(5.0)
        response = requests.get(url).json()['result']['items']
    
    # Время между запроса к API, выбарно имперически
    sleep(0.5)

    print('-------------------------------------------------')
    
    # Отбрасываем СИ поверенные до даты сдачи на поверку
    # response = get_actual_date(response, date_of_dispatch) if response else []
    
    # Если СИ не поверен, проверяем следующую запись
    if not response:
        print(mi_number, 'не поверен')
        continue

    print('Найдены совпадения:')
    print(response[0]['mit_title'], response[0]['mit_notation'],
          response[0]['mi_number'])
    print('=============================')
    for count, mi in enumerate(response):
        print(f"<{count}> {mi['org_title']}")

    print('=============================')
    print('Выберите номер поверенного СИ для внесения в базу')
    choised_mi = input('==>')

    
    if not choised_mi.isdigit():
        continue
    choised_mi = int(choised_mi)
    if int(choised_mi) not in tuple(range(len(response))):
        continue

    # Вносим данные в таблицу
    # Свидетельство 
    sheet.cell(row=row, column=16).value = response[choised_mi]['result_docnum']
    # Дата поверки
    sheet.cell(row=row, column=15).value = datetime.datetime.strptime(response[choised_mi]['verification_date'],'%d.%m.%Y')
    # Результат поверки
    if response[choised_mi]['applicability']:
        sheet.cell(row=row, column=14).value = 'свидетельство'
    else:
        sheet.cell(row=row, column=14).value = 'извещение'

    # Вывод результата в терминал
    mit_title = response[choised_mi]['mit_title']
    mi_number = response[choised_mi]['mi_number']
    mit_notation = response[choised_mi]['mit_notation']
    # print(f'Поверен ==> {mit_title} {mit_notation}, зав.номер {mi_number} ')
    # print('=======================================================')

# Сохраняем результаты в файле
print('Записать сведения о поверенных СИ в файл?')
answer = input('==>')

if answer.lower() in ('y', 'д', '1'):
    wb.save('check.xlsx')

wb.close()

